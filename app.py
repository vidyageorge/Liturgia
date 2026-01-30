from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_cors import CORS
from models import db, Member, Priest, CommunityRole, MassType, Mass, MassAssignment, Apostle, DepartedSoul, MassChangeLog
from datetime import datetime, date
import json
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///liturgia.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'st-mathias-english-community-2024'

db.init_app(app)

def init_db():
    """Initialize database with default mass types"""
    with app.app_context():
        db.create_all()
        
        # Check if mass types already exist
        if MassType.query.count() == 0:
            mass_types = [
                {
                    'name': 'Sunday Mass',
                    'default_time': '7:15 AM',
                    'description': 'Regular Sunday Mass',
                    'is_special_event': False,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'Christmas Mass',
                    'default_time': '',
                    'description': 'Christmas Celebration with Carols',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_carols': True
                },
                {
                    'name': 'New Year Mass',
                    'default_time': '11:15 PM',
                    'description': 'New Year Thanksgiving Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_thanksgiving': True
                },
                {
                    'name': 'Palm Sunday',
                    'default_time': '7:30 AM',
                    'description': 'Palm Sunday Mass with Gospel Narration',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_gospel_narration': True,
                    'has_mc_reader': True
                },
                {
                    'name': 'Maundy Thursday',
                    'default_time': '5:30 PM',
                    'description': 'Maundy Thursday with 12 Apostles',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_mc_reader': True,
                    'has_apostles': True
                },
                {
                    'name': 'Good Friday',
                    'default_time': '5:30 PM',
                    'description': 'Good Friday with Morning Adoration and Gospel Narration',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_gospel_narration': True,
                    'has_mc_reader': True,
                    'has_morning_adoration': True
                },
                {
                    'name': 'Easter Vigil',
                    'default_time': '11:15 PM',
                    'description': 'Easter Vigil Mass with Three Readings',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'third_reading', 'prayer_of_faithful']),
                    'has_mc_reader': True,
                    'has_third_reading': True
                },
                {
                    'name': 'All Souls Day',
                    'default_time': '7:15 AM',
                    'description': 'All Souls Day - Remembrance of Departed Souls',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful']),
                    'has_departed_souls_reader': True
                },
                {
                    'name': 'Independence Day Mass',
                    'default_time': '7:15 AM',
                    'description': 'Independence Day Special Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'Pongal Mass',
                    'default_time': '7:15 AM',
                    'description': 'Pongal Festival Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'St. Cecilia Feast',
                    'default_time': '7:15 AM',
                    'description': 'Feast of St. Cecilia - Patron Saint of Musicians',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'Christmas Morning Mass',
                    'default_time': '7:15 AM',
                    'description': 'Christmas Day Morning Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'New Year Morning Mass',
                    'default_time': '7:15 AM',
                    'description': 'New Year Day Morning Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                },
                {
                    'name': 'Easter Morning Mass',
                    'default_time': '7:15 AM',
                    'description': 'Easter Sunday Morning Mass',
                    'is_special_event': True,
                    'required_roles': json.dumps(['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])
                }
            ]
            
            for mt_data in mass_types:
                mt = MassType(**mt_data)
                db.session.add(mt)
            
            # Add default community roles
            roles = [
                {'name': 'Liturgy Coordinator', 'category': 'liturgy', 'description': 'Overall liturgy coordination'},
                {'name': 'Liturgy Member', 'category': 'liturgy', 'description': 'Liturgy team member'},
                {'name': 'Executive Member', 'category': 'executive', 'description': 'Executive committee member'},
                {'name': 'Choir Coordinator', 'category': 'choir', 'description': 'Choir incharge'},
                {'name': 'Choir Member', 'category': 'choir', 'description': 'Choir team member'},
                {'name': 'Catechism Coordinator', 'category': 'catechism', 'description': 'Catechism incharge'},
                {'name': 'Catechism Teacher', 'category': 'catechism', 'description': 'Catechism teacher'},
                {'name': 'Volunteer Coordinator', 'category': 'volunteers', 'description': 'Volunteers incharge'},
                {'name': 'Volunteer', 'category': 'volunteers', 'description': 'Community volunteer'}
            ]
            
            for role_data in roles:
                role = CommunityRole(**role_data)
                db.session.add(role)
            
            db.session.commit()

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/members')
def members_page():
    return render_template('members.html')

@app.route('/masses')
def masses_page():
    return render_template('masses.html')

@app.route('/community')
def community_page():
    return render_template('community.html')

@app.route('/schedule')
def schedule_page():
    return render_template('schedule.html')

@app.route('/changelog')
def changelog_page():
    return render_template('changelog.html')

# API Routes - Members
@app.route('/api/members', methods=['GET'])
def get_members():
    members = Member.query.filter_by(active=True).order_by(Member.name).all()
    
    # Get current year for reading count
    current_year = datetime.now().year
    
    result = []
    for m in members:
        member_dict = m.to_dict()
        # Count readings for current year using SQLite strftime
        from sqlalchemy import func
        reading_count = MassAssignment.query.join(Mass).filter(
            MassAssignment.member_id == m.id,
            func.strftime('%Y', Mass.date) == str(current_year)
        ).count()
        # Total readings (all time)
        total_readings = MassAssignment.query.filter_by(member_id=m.id).count()
        
        # Use total_readings as the main count (since we're looking at all history)
        member_dict['reading_count'] = total_readings  # Show total, not just current year
        member_dict['current_year_count'] = reading_count  # Add current year separately
        member_dict['total_readings'] = total_readings
        member_dict['created_at'] = m.created_at.strftime('%Y-%m-%d') if m.created_at else None
        member_dict['member_since'] = m.created_at.strftime('%b %Y') if m.created_at else 'Unknown'
        member_dict['experience_level'] = m.experience_level
        member_dict['years_of_service'] = m.years_of_service
        result.append(member_dict)
    
    return jsonify(result)

@app.route('/api/members', methods=['POST'])
def create_member():
    data = request.json
    member = Member(
        name=data['name'],
        phone=data.get('phone'),
        email=data.get('email')
    )
    db.session.add(member)
    db.session.commit()
    return jsonify(member.to_dict()), 201

@app.route('/api/members/<int:id>', methods=['GET'])
def get_member(id):
    member = Member.query.get_or_404(id)
    return jsonify(member.to_dict())

@app.route('/api/members/<int:id>', methods=['PUT'])
def update_member(id):
    member = Member.query.get_or_404(id)
    data = request.json
    member.name = data.get('name', member.name)
    member.phone = data.get('phone', member.phone)
    member.email = data.get('email', member.email)
    member.active = data.get('active', member.active)
    member.experience_level = data.get('experience_level', member.experience_level)
    member.years_of_service = data.get('years_of_service', member.years_of_service)
    db.session.commit()
    return jsonify(member.to_dict())

@app.route('/api/members/<int:id>', methods=['DELETE'])
def delete_member(id):
    member = Member.query.get_or_404(id)
    member.active = False  # Soft delete
    db.session.commit()
    return jsonify({'message': 'Member deactivated'})

@app.route('/api/members/<int:id>/roles', methods=['POST'])
def assign_member_role(id):
    member = Member.query.get_or_404(id)
    data = request.json
    role = CommunityRole.query.get_or_404(data['role_id'])
    if role not in member.roles:
        member.roles.append(role)
        db.session.commit()
    return jsonify(member.to_dict())

@app.route('/api/members/<int:id>/roles/<int:role_id>', methods=['DELETE'])
def remove_member_role(id, role_id):
    member = Member.query.get_or_404(id)
    role = CommunityRole.query.get_or_404(role_id)
    if role in member.roles:
        member.roles.remove(role)
        db.session.commit()
    return jsonify(member.to_dict())

# API Routes - Priests
@app.route('/api/priests', methods=['GET'])
def get_priests():
    priests = Priest.query.filter_by(active=True).all()
    return jsonify([p.to_dict() for p in priests])

@app.route('/api/priests', methods=['POST'])
def create_priest():
    data = request.json
    priest = Priest(
        name=data['name'],
        title=data.get('title', 'Fr.'),
        phone=data.get('phone')
    )
    db.session.add(priest)
    db.session.commit()
    return jsonify(priest.to_dict()), 201

@app.route('/api/priests/<int:id>', methods=['PUT'])
def update_priest(id):
    priest = Priest.query.get_or_404(id)
    data = request.json
    priest.name = data.get('name', priest.name)
    priest.title = data.get('title', priest.title)
    priest.phone = data.get('phone', priest.phone)
    priest.active = data.get('active', priest.active)
    db.session.commit()
    return jsonify(priest.to_dict())

@app.route('/api/priests/<int:id>', methods=['DELETE'])
def delete_priest(id):
    priest = Priest.query.get_or_404(id)
    priest.active = False
    db.session.commit()
    return jsonify({'message': 'Priest deactivated'})

# API Routes - Community Roles
@app.route('/api/roles', methods=['GET'])
def get_roles():
    roles = CommunityRole.query.all()
    return jsonify([r.to_dict() for r in roles])

@app.route('/api/roles', methods=['POST'])
def create_role():
    data = request.json
    role = CommunityRole(
        name=data['name'],
        category=data.get('category'),
        description=data.get('description')
    )
    db.session.add(role)
    db.session.commit()
    return jsonify(role.to_dict()), 201

# API Routes - Mass Types
@app.route('/api/mass-types', methods=['GET'])
def get_mass_types():
    mass_types = MassType.query.all()
    return jsonify([mt.to_dict() for mt in mass_types])

@app.route('/api/mass-types', methods=['POST'])
def create_mass_type():
    data = request.json
    mass_type = MassType(
        name=data['name'],
        default_time=data.get('default_time'),
        description=data.get('description'),
        is_special_event=data.get('is_special_event', True),
        required_roles=json.dumps(data.get('required_roles', ['introduction', 'first_reading', 'second_reading', 'prayer_of_faithful'])),
        has_gospel_narration=data.get('has_gospel_narration', False),
        has_mc_reader=data.get('has_mc_reader', False),
        has_third_reading=data.get('has_third_reading', False),
        has_apostles=data.get('has_apostles', False),
        has_thanksgiving=data.get('has_thanksgiving', False),
        has_carols=data.get('has_carols', False),
        has_morning_adoration=data.get('has_morning_adoration', False),
        has_departed_souls_reader=data.get('has_departed_souls_reader', False)
    )
    db.session.add(mass_type)
    db.session.commit()
    return jsonify(mass_type.to_dict()), 201

# API Routes - Masses
@app.route('/api/masses', methods=['GET'])
def get_masses():
    masses = Mass.query.order_by(Mass.date.desc()).all()
    return jsonify([m.to_dict() for m in masses])

@app.route('/api/masses/upcoming', methods=['GET'])
def get_upcoming_masses():
    today = date.today()
    masses = Mass.query.filter(Mass.date >= today).order_by(Mass.date.asc()).all()
    return jsonify([m.to_dict() for m in masses])

@app.route('/api/masses/past', methods=['GET'])
def get_past_masses():
    today = date.today()
    masses = Mass.query.filter(Mass.date < today).order_by(Mass.date.desc()).limit(50).all()
    return jsonify([m.to_dict() for m in masses])

@app.route('/api/masses', methods=['POST'])
def create_mass():
    data = request.json
    mass = Mass(
        mass_type_id=data['mass_type_id'],
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        time=data.get('time'),
        celebrant=data.get('celebrant'),
        notes=data.get('notes')
    )
    db.session.add(mass)
    db.session.commit()
    return jsonify(mass.to_dict()), 201

@app.route('/api/masses/<int:id>', methods=['GET'])
def get_mass(id):
    mass = Mass.query.get_or_404(id)
    return jsonify(mass.to_dict())

@app.route('/api/masses/<int:id>', methods=['PUT'])
def update_mass(id):
    mass = Mass.query.get_or_404(id)
    data = request.json
    
    # Require change reason
    change_reason = data.get('change_reason')
    if not change_reason or not change_reason.strip():
        return jsonify({'error': 'Change reason is required'}), 400
    
    # Track changes
    changes = []
    
    # Check mass_type_id change
    if 'mass_type_id' in data and data['mass_type_id'] != mass.mass_type_id:
        old_type = mass.mass_type.name if mass.mass_type else 'None'
        new_type_obj = MassType.query.get(data['mass_type_id'])
        new_type = new_type_obj.name if new_type_obj else 'None'
        
        changes.append({
            'field': 'mass_type',
            'old': old_type,
            'new': new_type
        })
        mass.mass_type_id = data['mass_type_id']
    
    # Check date change
    if 'date' in data:
        new_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if new_date != mass.date:
            changes.append({
                'field': 'date',
                'old': str(mass.date),
                'new': str(new_date)
            })
            mass.date = new_date
    
    # Check time change
    if 'time' in data and data['time'] != mass.time:
        changes.append({
            'field': 'time',
            'old': mass.time or '',
            'new': data['time']
        })
        mass.time = data['time']
    
    # Check celebrant change
    if 'celebrant' in data:
        old_celebrant = mass.celebrant or ''
        new_celebrant = data['celebrant'] or ''
        if old_celebrant != new_celebrant:
            changes.append({
                'field': 'celebrant',
                'old': old_celebrant,
                'new': new_celebrant
            })
        mass.celebrant = data['celebrant']
    
    # Check notes change
    if 'notes' in data and data['notes'] != mass.notes:
        changes.append({
            'field': 'notes',
            'old': mass.notes or '',
            'new': data['notes']
        })
        mass.notes = data['notes']
    
    # Log all changes
    for change in changes:
        log_entry = MassChangeLog(
            mass_id=mass.id,
            changed_by=data.get('changed_by', 'System'),
            change_reason=change_reason,
            field_changed=change['field'],
            old_value=change['old'],
            new_value=change['new']
        )
        db.session.add(log_entry)
    
    db.session.commit()
    return jsonify(mass.to_dict())

@app.route('/api/masses/<int:id>', methods=['DELETE'])
def delete_mass(id):
    mass = Mass.query.get_or_404(id)
    db.session.delete(mass)
    db.session.commit()
    return jsonify({'message': 'Mass deleted'})

# API Routes - Mass Assignments
@app.route('/api/masses/<int:mass_id>/assignments', methods=['POST'])
def create_assignment(mass_id):
    mass = Mass.query.get_or_404(mass_id)
    data = request.json
    
    # Check if this exact assignment already exists (same role AND same member/name)
    member_id = data.get('member_id')
    member_name_override = data.get('member_name_override')
    
    existing = None
    if member_id:
        existing = MassAssignment.query.filter_by(
            mass_id=mass_id, 
            role=data['role'], 
            member_id=member_id
        ).first()
    elif member_name_override:
        existing = MassAssignment.query.filter_by(
            mass_id=mass_id, 
            role=data['role'], 
            member_name_override=member_name_override
        ).first()
    
    if existing:
        # Update existing assignment
        existing.member_id = member_id
        existing.member_name_override = member_name_override
        existing.notes = data.get('notes')
    else:
        # Create new assignment (allows multiple readers for same role)
        assignment = MassAssignment(
            mass_id=mass_id,
            member_id=member_id,
            role=data['role'],
            member_name_override=member_name_override,
            notes=data.get('notes')
        )
        db.session.add(assignment)
    
    db.session.commit()
    return jsonify(mass.to_dict())

@app.route('/api/masses/<int:mass_id>/assignments/<int:id>', methods=['DELETE'])
def delete_assignment(mass_id, id):
    assignment = MassAssignment.query.get_or_404(id)
    db.session.delete(assignment)
    db.session.commit()
    return jsonify({'message': 'Assignment deleted'})

# Bulk update assignments with change logging
@app.route('/api/masses/<int:mass_id>/assignments/bulk-update', methods=['POST'])
def bulk_update_assignments(mass_id):
    mass = Mass.query.get_or_404(mass_id)
    data = request.json
    new_assignments = data.get('assignments', [])
    change_reason = data.get('change_reason')
    changed_by = data.get('changed_by', 'System')
    
    # Get existing assignments
    existing_assignments = MassAssignment.query.filter_by(mass_id=mass_id).all()
    
    # Create a map of existing assignments by role
    existing_by_role = {}
    for a in existing_assignments:
        role = a.role
        if role not in existing_by_role:
            existing_by_role[role] = []
        reader_name = a.member.name if a.member else a.member_name_override
        existing_by_role[role].append(reader_name)
    
    # Create a map of new assignments by role
    new_by_role = {}
    for a in new_assignments:
        role = a['role']
        if role not in new_by_role:
            new_by_role[role] = []
        new_by_role[role].append(a['member_name'])
    
    # Compare and log changes
    all_roles = set(list(existing_by_role.keys()) + list(new_by_role.keys()))
    for role in all_roles:
        old_readers = sorted(existing_by_role.get(role, []))
        new_readers = sorted(new_by_role.get(role, []))
        
        if old_readers != new_readers:
            # Log the change
            log_entry = MassChangeLog(
                mass_id=mass_id,
                changed_by=changed_by,
                change_reason=change_reason or 'Reader assignment updated',
                field_changed=f'{role}_reader',
                old_value=', '.join(old_readers) if old_readers else '(none)',
                new_value=', '.join(new_readers) if new_readers else '(none)'
            )
            db.session.add(log_entry)
    
    # Delete all existing assignments
    for a in existing_assignments:
        db.session.delete(a)
    
    # Create new assignments
    for a in new_assignments:
        assignment = MassAssignment(
            mass_id=mass_id,
            member_id=a.get('member_id'),
            role=a['role'],
            member_name_override=a.get('member_name') if not a.get('member_id') else None,
            notes=a.get('notes')
        )
        db.session.add(assignment)
    
    db.session.commit()
    return jsonify(mass.to_dict())

# API Routes - Apostles (for Maundy Thursday)
@app.route('/api/masses/<int:mass_id>/apostles', methods=['POST'])
def add_apostle(mass_id):
    mass = Mass.query.get_or_404(mass_id)
    data = request.json
    
    apostle = Apostle(
        mass_id=mass_id,
        apostle_name=data['apostle_name'],
        member_id=data.get('member_id'),
        member_name_override=data.get('member_name_override')
    )
    db.session.add(apostle)
    db.session.commit()
    return jsonify(mass.to_dict())

@app.route('/api/masses/<int:mass_id>/apostles/<int:id>', methods=['DELETE'])
def delete_apostle(mass_id, id):
    apostle = Apostle.query.get_or_404(id)
    db.session.delete(apostle)
    db.session.commit()
    return jsonify({'message': 'Apostle removed'})

# API Routes - Departed Souls (for All Souls Day)
@app.route('/api/masses/<int:mass_id>/departed-souls', methods=['GET'])
def get_departed_souls(mass_id):
    souls = DepartedSoul.query.filter_by(mass_id=mass_id).all()
    return jsonify([s.to_dict() for s in souls])

@app.route('/api/masses/<int:mass_id>/departed-souls', methods=['POST'])
def add_departed_soul(mass_id):
    mass = Mass.query.get_or_404(mass_id)
    data = request.json
    
    soul = DepartedSoul(
        mass_id=mass_id,
        name=data['name'],
        family_name=data.get('family_name')
    )
    db.session.add(soul)
    db.session.commit()
    return jsonify(soul.to_dict()), 201

@app.route('/api/masses/<int:mass_id>/departed-souls/<int:id>', methods=['DELETE'])
def delete_departed_soul(mass_id, id):
    soul = DepartedSoul.query.get_or_404(id)
    db.session.delete(soul)
    db.session.commit()
    return jsonify({'message': 'Departed soul removed'})

# Statistics API
@app.route('/api/stats', methods=['GET'])
def get_stats():
    from sqlalchemy import func
    
    total_members = Member.query.filter_by(active=True).count()
    total_masses = Mass.query.count()
    upcoming_masses = Mass.query.filter(Mass.date >= date.today()).count()
    
    # Get top 3 volunteers of all time
    top_volunteers = db.session.query(
        Member.id,
        Member.name,
        func.count(MassAssignment.id).label('reading_count')
    ).join(
        MassAssignment, Member.id == MassAssignment.member_id
    ).filter(
        Member.active == True
    ).group_by(
        Member.id, Member.name
    ).order_by(
        func.count(MassAssignment.id).desc()
    ).limit(3).all()
    
    top_volunteers_data = [
        {
            'id': v.id,
            'name': v.name,
            'reading_count': v.reading_count
        }
        for v in top_volunteers
    ]
    
    return jsonify({
        'total_members': total_members,
        'total_masses': total_masses,
        'upcoming_masses': upcoming_masses,
        'top_volunteers': top_volunteers_data
    })

# Member reading history
@app.route('/api/members/<int:id>/history', methods=['GET'])
def get_member_history(id):
    member = Member.query.get_or_404(id)
    assignments = MassAssignment.query.filter_by(member_id=id).all()
    history = []
    for a in assignments:
        if a.mass:
            history.append({
                'date': a.mass.date.strftime('%Y-%m-%d'),
                'mass_type': a.mass.mass_type.name if a.mass.mass_type else 'Unknown',
                'role': a.role
            })
    return jsonify(history)

# Export mass history to Excel
@app.route('/api/masses/export', methods=['GET'])
def export_masses():
    """Export all masses to Excel format"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    year = request.args.get('year')
    
    # Query masses
    query = Mass.query.order_by(Mass.date.asc())
    
    # Filter by year if specified
    if year:
        query = query.filter(db.extract('year', Mass.date) == int(year))
    # Otherwise filter by date range
    elif start_date:
        query = query.filter(Mass.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Mass.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    masses = query.all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mass History {year}" if year else "Mass History"
    
    # Define headers
    headers = ['Date', 'Day', 'Mass Type', 'Time', 'Celebrant', 
               'Introduction', 'First Reading', 'Second Reading', 'Prayer of Faithful', 
               'MC Reader', 'Third Reading', 'Gospel Narrators', 'Notes']
    
    # Style for header
    header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, mass in enumerate(masses, 2):
        # Get all assignments for this mass
        assignments = {a.role: a.member.name if a.member else a.member_name_override 
                      for a in mass.assignments}
        
        # Date
        ws.cell(row=row_idx, column=1, value=mass.date.strftime('%Y-%m-%d'))
        
        # Day of week
        day_name = mass.date.strftime('%A')
        ws.cell(row=row_idx, column=2, value=day_name)
        
        # Mass Type
        ws.cell(row=row_idx, column=3, value=mass.mass_type.name if mass.mass_type else '')
        
        # Time
        ws.cell(row=row_idx, column=4, value=mass.time or '')
        
        # Celebrant
        ws.cell(row=row_idx, column=5, value=mass.celebrant or '')
        
        # Introduction
        ws.cell(row=row_idx, column=6, value=assignments.get('introduction', ''))
        
        # First Reading
        ws.cell(row=row_idx, column=7, value=assignments.get('first_reading', ''))
        
        # Second Reading
        ws.cell(row=row_idx, column=8, value=assignments.get('second_reading', ''))
        
        # Prayer of Faithful
        ws.cell(row=row_idx, column=9, value=assignments.get('prayer_of_faithful', ''))
        
        # MC Reader
        ws.cell(row=row_idx, column=10, value=assignments.get('mc_reader', ''))
        
        # Third Reading
        ws.cell(row=row_idx, column=11, value=assignments.get('third_reading', ''))
        
        # Gospel Narrators (if multiple)
        gospel_narrators = [v for k, v in assignments.items() if 'gospel_narrator' in k]
        ws.cell(row=row_idx, column=12, value=', '.join(gospel_narrators) if gospel_narrators else '')
        
        # Notes
        ws.cell(row=row_idx, column=13, value=mass.notes or '')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date range or year
    if year:
        filename = f'mass_history_{year}.xlsx'
    elif start_date and end_date:
        filename = f'mass_history_{start_date}_to_{end_date}.xlsx'
    else:
        filename = f'mass_history_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Get change history for a mass
@app.route('/api/masses/<int:id>/history', methods=['GET'])
def get_mass_history(id):
    mass = Mass.query.get_or_404(id)
    logs = MassChangeLog.query.filter_by(mass_id=id).order_by(MassChangeLog.changed_at.desc()).all()
    return jsonify([log.to_dict() for log in logs])

# Get all changes (changelog page)
@app.route('/api/changelog', methods=['GET'])
def get_all_changes():
    limit = request.args.get('limit', type=int)
    
    query = MassChangeLog.query.order_by(MassChangeLog.changed_at.desc())
    
    if limit:
        query = query.limit(limit)
    
    logs = query.all()
    
    # Enrich with mass information
    result = []
    for log in logs:
        log_dict = log.to_dict()
        mass = Mass.query.get(log.mass_id)
        if mass:
            log_dict['mass_date'] = mass.date.strftime('%Y-%m-%d')
            log_dict['mass_type'] = mass.mass_type.name if mass.mass_type else 'Unknown'
        else:
            log_dict['mass_date'] = None
            log_dict['mass_type'] = 'Unknown'
        result.append(log_dict)
    
    return jsonify(result)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)

